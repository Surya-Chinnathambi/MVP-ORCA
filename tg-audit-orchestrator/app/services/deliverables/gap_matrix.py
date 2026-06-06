"""Gap matrix deliverable — requirements × coverage × evidence × findings.

Exports to XLSX (openpyxl) and HTML. Creates/versions a Deliverable record.
"""
from __future__ import annotations

import uuid
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.styles import Font, PatternFill
from sqlalchemy.orm import Session

from app.models.clients import Project
from app.models.delivery import Deliverable, DeliverableKind
from app.models.evidence import EvidenceItem, EvidenceRequest
from app.models.scope import Requirement
from app.models.tasks import Finding, FindingSeverity


_SEVERITY_ORDER = {
    FindingSeverity.critical.value: 0,
    FindingSeverity.high.value: 1,
    FindingSeverity.medium.value: 2,
    FindingSeverity.low.value: 3,
    FindingSeverity.info.value: 4,
}

_SEVERITY_COLOURS = {
    "critical": "C00000",
    "high": "FF0000",
    "medium": "FF9900",
    "low": "FFFF00",
    "info": "CCCCCC",
}


def _next_version(db: Session, project_id: str, kind: DeliverableKind) -> int:
    existing = (
        db.query(Deliverable)
        .filter_by(project_id=project_id, kind=kind)
        .order_by(Deliverable.version.desc())
        .first()
    )
    return (existing.version + 1) if existing else 1


def _build_matrix_rows(db: Session, project_id: str) -> list[dict]:
    """Return one row per requirement with linked evidence and findings."""
    requirements = (
        db.query(Requirement).filter_by(project_id=project_id).all()
    )
    ev_requests = (
        db.query(EvidenceRequest).filter_by(project_id=project_id).all()
    )
    findings = db.query(Finding).filter_by(project_id=project_id).all()

    req_to_ev: dict[str, list[EvidenceRequest]] = {}
    for er in ev_requests:
        if er.requirement_id:
            req_to_ev.setdefault(er.requirement_id, []).append(er)

    req_to_findings: dict[str, list[Finding]] = {}
    for f in findings:
        if f.requirement_id:
            req_to_findings.setdefault(f.requirement_id, []).append(f)

    rows = []
    for req in requirements:
        linked_ev = req_to_ev.get(req.id, [])
        linked_findings = req_to_findings.get(req.id, [])

        received_count = sum(
            1 for er in linked_ev if er.status in ("received", "waived")
        )
        coverage = "covered" if received_count > 0 else (
            "partial" if linked_ev else "gap"
        )
        worst_severity = None
        if linked_findings:
            sorted_f = sorted(
                linked_findings,
                key=lambda f: _SEVERITY_ORDER.get(f.severity, 99),
            )
            worst_severity = sorted_f[0].severity

        rows.append({
            "ref_code": req.ref_code,
            "category": req.category,
            "text": req.text,
            "evidence_expectation": req.evidence_expectation or "",
            "evidence_requests": len(linked_ev),
            "evidence_received": received_count,
            "coverage": coverage,
            "findings_count": len(linked_findings),
            "worst_severity": worst_severity or "",
        })
    return rows


def generate_gap_matrix(
    db: Session,
    project: Project,
    output_dir: Path,
    actor_id: Optional[str] = None,
) -> tuple[Deliverable, Deliverable]:
    """Generate XLSX + HTML gap matrices. Returns (xlsx_deliverable, html_deliverable)."""
    output_dir.mkdir(parents=True, exist_ok=True)
    rows = _build_matrix_rows(db, project.id)
    now = datetime.now(timezone.utc)

    xlsx_version = _next_version(db, project.id, DeliverableKind.gap_matrix)
    xlsx_path = output_dir / f"gap_matrix_v{xlsx_version}.xlsx"
    _write_xlsx(rows, xlsx_path)

    html_path = output_dir / f"gap_matrix_v{xlsx_version}.html"
    _write_html(rows, html_path, project)

    xlsx_del = Deliverable(
        id=str(uuid.uuid4()),
        project_id=project.id,
        kind=DeliverableKind.gap_matrix,
        format="xlsx",
        file_path=str(xlsx_path),
        generated_at=now,
        version=xlsx_version,
    )
    html_del = Deliverable(
        id=str(uuid.uuid4()),
        project_id=project.id,
        kind=DeliverableKind.gap_matrix,
        format="html",
        file_path=str(html_path),
        generated_at=now,
        version=xlsx_version,
    )
    db.add(xlsx_del)
    db.add(html_del)
    return xlsx_del, html_del


def _write_xlsx(rows: list[dict], path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Gap Matrix"

    headers = [
        "Ref Code", "Category", "Requirement", "Evidence Expectation",
        "Evidence Requests", "Evidence Received", "Coverage",
        "Findings Count", "Worst Severity",
    ]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E79")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill

    for row_idx, row in enumerate(rows, 2):
        values = [
            row["ref_code"], row["category"], row["text"],
            row["evidence_expectation"], row["evidence_requests"],
            row["evidence_received"], row["coverage"],
            row["findings_count"], row["worst_severity"],
        ]
        for col, val in enumerate(values, 1):
            ws.cell(row=row_idx, column=col, value=val)

        sev = row["worst_severity"]
        if sev in _SEVERITY_COLOURS:
            fill = PatternFill("solid", fgColor=_SEVERITY_COLOURS[sev])
            ws.cell(row=row_idx, column=9).fill = fill

    for col in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

    wb.save(path)


def _write_html(rows: list[dict], path: Path, project: Project) -> None:
    rows_html = ""
    for row in rows:
        cov_colour = {"covered": "#c6efce", "partial": "#ffeb9c", "gap": "#ffc7ce"}.get(
            row["coverage"], "#ffffff"
        )
        rows_html += (
            f'<tr>'
            f'<td>{row["ref_code"]}</td>'
            f'<td>{row["category"]}</td>'
            f'<td>{row["text"]}</td>'
            f'<td>{row["evidence_expectation"]}</td>'
            f'<td>{row["evidence_requests"]}</td>'
            f'<td>{row["evidence_received"]}</td>'
            f'<td style="background:{cov_colour}">{row["coverage"]}</td>'
            f'<td>{row["findings_count"]}</td>'
            f'<td>{row["worst_severity"]}</td>'
            f'</tr>\n'
        )

    html = f"""<!DOCTYPE html>
<html><head><meta charset="utf-8">
<title>Gap Matrix — {project.id}</title>
<style>
body{{font-family:Arial,sans-serif;font-size:13px;margin:20px}}
table{{border-collapse:collapse;width:100%}}
th,td{{border:1px solid #ccc;padding:6px 10px;text-align:left}}
th{{background:#1f4e79;color:#fff}}
tr:nth-child(even){{background:#f5f5f5}}
</style></head>
<body>
<h2>Gap Matrix</h2>
<p>Project: {project.id} | Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}</p>
<table>
<thead><tr>
<th>Ref Code</th><th>Category</th><th>Requirement</th><th>Evidence Expectation</th>
<th>Ev Requests</th><th>Ev Received</th><th>Coverage</th>
<th>Findings</th><th>Worst Severity</th>
</tr></thead>
<tbody>{rows_html}</tbody>
</table>
</body></html>"""
    path.write_text(html, encoding="utf-8")

"""Stage 10 acceptance test — Deliverable Builder.

Verifies:
1. Gap matrix XLSX generated with correct rows.
2. Roadmap Markdown + HTML generated.
3. Report HTML generated as draft.
4. Report cannot be released without an approved ApprovalRequest (Gate 6).
5. Deliverable.version increments on regeneration.
"""
import tempfile
from pathlib import Path

import pytest
from sqlalchemy import create_engine
from sqlalchemy.orm import Session, sessionmaker
from sqlalchemy.pool import StaticPool

import app.models  # noqa: F401 — ensures all mappers register
from app.db import Base
from app.models.clients import Client, Project, ServiceType
from app.models.delivery import Deliverable, DeliverableKind, RemediationAction
from app.models.evidence import EvidenceItem, EvidenceRequest, EvidenceRequestStatus, ReviewerStatus
from app.models.scope import Requirement
from app.models.tasks import Finding, FindingSeverity, FindingSource, FindingStatus
from app.models.users import Role, RoleName, User
from app.models.workflow import ApprovalRequest, ApprovalStatus
from app.services.auth import hash_password
from app.services.audit import decide_approval, record_event, request_approval
from app.services.deliverables.gap_matrix import generate_gap_matrix
from app.services.deliverables.report import generate_report, has_release_approval
from app.services.deliverables.roadmap import generate_roadmap


# ── Fixtures ──────────────────────────────────────────────────────────────────

@pytest.fixture(scope="module")
def engine():
    e = create_engine(
        "sqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(e)
    yield e
    Base.metadata.drop_all(e)


@pytest.fixture(scope="module")
def SessionTest(engine):
    return sessionmaker(bind=engine, autocommit=False, autoflush=False)


@pytest.fixture(scope="module")
def seed(engine, SessionTest):
    """Create roles, admin user, client, project, requirements, evidence, findings."""
    with Session(engine) as db:
        for name in [r.value for r in RoleName]:
            db.add(Role(name=name))

        admin = User(
            email="admin@test.local",
            password_hash=hash_password("pass"),
            full_name="Admin",
            is_active=True,
        )
        db.add(admin)
        db.flush()

        client = Client(entity_name="Acme Corp")
        db.add(client)
        db.flush()

        project = Project(
            client_id=client.id,
            service_type=ServiceType.dpdp,
            owner_id=admin.id,
            pack_id="dpdp",
            gates={},
        )
        db.add(project)
        db.flush()

        # Two requirements
        req1 = Requirement(
            project_id=project.id, ref_code="DPDP-01",
            text="Privacy notice requirement", category="notice",
            evidence_expectation="Current privacy notice document",
        )
        req2 = Requirement(
            project_id=project.id, ref_code="DPDP-02",
            text="Consent collection", category="consent",
            evidence_expectation="Consent form samples",
        )
        db.add_all([req1, req2])
        db.flush()

        # Evidence request linked to req1 (received) and req2 (open → gap)
        er1 = EvidenceRequest(
            project_id=project.id, requirement_id=req1.id,
            title="Privacy notice document", status=EvidenceRequestStatus.received,
        )
        er2 = EvidenceRequest(
            project_id=project.id, requirement_id=req2.id,
            title="Consent form", status=EvidenceRequestStatus.open,
        )
        db.add_all([er1, er2])
        db.flush()

        # One evidence item for req1
        ev = EvidenceItem(
            project_id=project.id,
            evidence_request_id=er1.id,
            source_file="privacy_notice.pdf",
            sha256="a" * 64,
            mime="application/pdf",
            classification="notice",
            reviewer_status=ReviewerStatus.accepted,
        )
        db.add(ev)
        db.flush()

        # One finding linked to req2
        finding = Finding(
            project_id=project.id, requirement_id=req2.id,
            title="No consent mechanism",
            description="Consent is not collected prior to data processing.",
            severity=FindingSeverity.high,
            status=FindingStatus.open,
            source=FindingSource.manual,
            evidence_item_ids=[ev.id],
        )
        db.add(finding)
        db.flush()

        # Remediation action for the finding
        action = RemediationAction(
            finding_id=finding.id,
            project_id=project.id,
            action="Implement consent collection flow before data processing.",
            status="open",
        )
        db.add(action)
        db.commit()

        return {
            "project_id": project.id,
            "admin_id": admin.id,
            "req1_id": req1.id,
            "req2_id": req2.id,
            "finding_id": finding.id,
            "ev_id": ev.id,
        }


# ── Tests ─────────────────────────────────────────────────────────────────────

def test_gap_matrix_xlsx_created(engine, SessionTest, seed, tmp_path):
    with Session(engine) as db:
        project = db.get(Project, seed["project_id"])
        xlsx_del, html_del = generate_gap_matrix(db, project, tmp_path / "gm")
        db.flush()
        xlsx_fmt = xlsx_del.format
        html_fmt = html_del.format
        xlsx_path = Path(xlsx_del.file_path)
        db.commit()

    assert xlsx_fmt == "xlsx"
    assert html_fmt == "html"
    assert xlsx_path.exists(), f"XLSX not found at {xlsx_path}"

    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    # Header row + 2 requirement rows
    assert ws.max_row == 3, f"Expected 3 rows (header + 2 reqs), got {ws.max_row}"


def test_gap_matrix_coverage_values(engine, SessionTest, seed, tmp_path):
    with Session(engine) as db:
        project = db.get(Project, seed["project_id"])
        xlsx_del, _ = generate_gap_matrix(db, project, tmp_path / "gm2")
        db.flush()
        xlsx_path = Path(xlsx_del.file_path)
        db.commit()

    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    # Row 2 = DPDP-01 (received → covered), row 3 = DPDP-02 (open request but none received → partial)
    coverage_col = 7  # column G
    row2_cov = ws.cell(row=2, column=coverage_col).value
    row3_cov = ws.cell(row=3, column=coverage_col).value
    assert row2_cov == "covered", f"Expected 'covered', got {row2_cov!r}"
    assert row3_cov == "partial", f"Expected 'partial', got {row3_cov!r}"


def test_gap_matrix_version_increments(engine, SessionTest, seed, tmp_path):
    with Session(engine) as db:
        project = db.get(Project, seed["project_id"])
        d1, _ = generate_gap_matrix(db, project, tmp_path / "gm3")
        db.flush()
        v1 = d1.version
        db.commit()

    with Session(engine) as db:
        project = db.get(Project, seed["project_id"])
        d2, _ = generate_gap_matrix(db, project, tmp_path / "gm3")
        db.flush()
        v2 = d2.version
        db.commit()

    assert v2 == v1 + 1, f"Version should increment: {v1} → {v2}"


def test_roadmap_markdown_created(engine, SessionTest, seed, tmp_path):
    with Session(engine) as db:
        project = db.get(Project, seed["project_id"])
        md_del, html_del = generate_roadmap(db, project, tmp_path / "rm")
        db.flush()
        md_fmt = md_del.format
        md_path = Path(md_del.file_path)
        db.commit()

    assert md_fmt == "md"
    assert md_path.exists()
    content = md_path.read_text()
    assert "# Remediation Roadmap" in content
    assert "HIGH" in content
    assert "Implement consent collection" in content


def test_roadmap_html_created(engine, SessionTest, seed, tmp_path):
    with Session(engine) as db:
        project = db.get(Project, seed["project_id"])
        _, html_del = generate_roadmap(db, project, tmp_path / "rm2")
        db.flush()
        html_path = Path(html_del.file_path)
        db.commit()

    assert html_path.exists()
    content = html_path.read_text()
    assert "Remediation Roadmap" in content
    assert "No consent" in content


def test_report_draft_generated(engine, SessionTest, seed, tmp_path):
    with Session(engine) as db:
        project = db.get(Project, seed["project_id"])
        deliverable = generate_report(db, project, tmp_path / "rpt")
        db.flush()
        fmt = deliverable.format
        html_path = Path(deliverable.file_path)
        db.commit()

    assert fmt == "html"
    assert html_path.exists()
    content = html_path.read_text()
    assert "DRAFT" in content
    assert "No consent mechanism" in content
    assert "privacy_notice.pdf" in content


def test_report_cannot_release_without_approval(engine, SessionTest, seed, tmp_path):
    """has_release_approval returns False until an approved ApprovalRequest exists."""
    with Session(engine) as db:
        project = db.get(Project, seed["project_id"])
        deliverable = generate_report(db, project, tmp_path / "rpt2")
        db.flush()
        del_id = deliverable.id
        db.commit()

    with Session(engine) as db:
        # No approval yet
        assert not has_release_approval(db, del_id)


def test_report_release_with_approval(engine, SessionTest, seed, tmp_path):
    """After an approved ApprovalRequest, has_release_approval returns True."""
    with Session(engine) as db:
        project = db.get(Project, seed["project_id"])
        deliverable = generate_report(db, project, tmp_path / "rpt3")
        db.flush()
        del_id = deliverable.id
        db.commit()

    with Session(engine) as db:
        approval = request_approval(
            db,
            project_id=seed["project_id"],
            target_type="deliverable",
            target_id=del_id,
            reason="Final report approved for release",
            approver_role="partner",
            requested_by=seed["admin_id"],
        )
        db.commit()
        approval_id = approval.id

    with Session(engine) as db:
        decide_approval(
            db,
            approval_id=approval_id,
            approved=True,
            decider_id=seed["admin_id"],
            reason="Approved",
        )
        db.commit()

    with Session(engine) as db:
        assert has_release_approval(db, del_id), "Should be releasable after approval"


def test_deliverables_list_populated(engine, SessionTest, seed):
    """All generated deliverables are queryable for the project."""
    with Session(engine) as db:
        items = (
            db.query(Deliverable)
            .filter_by(project_id=seed["project_id"])
            .all()
        )
    kinds = {d.kind for d in items}
    # gap_matrix, roadmap, and report were all generated in this test module
    assert DeliverableKind.gap_matrix in kinds
    assert DeliverableKind.roadmap in kinds
    assert DeliverableKind.report in kinds
